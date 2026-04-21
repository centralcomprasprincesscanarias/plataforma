"""
watcher_dossier.py  -  Princess Hotels
Vigila el Excel DOSSIER MENAJE y sincroniza con Supabase
cuando detecta un cambio en el fichero.

Instalar dependencias (una vez):
    py -m pip install watchdog openpyxl requests

Arrancar:
    py watcher_dossier.py
"""

import re, sys, time, logging, hashlib
import openpyxl, requests
from pathlib import Path
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ── CONFIG ─────────────────────────────────────────────────
EXCEL_FOLDER = r"T:\maspalomas&tabaibac\Compras-Chef\DOSSIER MENAJE"
EXCEL_PREFIX = "ACTIVO OPERACIONAL PRINCESS"
SHEET_NAME   = "DEPARTAMENTOS"

SUPABASE_URL = "https://ifiwuloipapsezmssyda.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImlmaXd1bG9pcGFwc2V6bXNzeWRhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzU4MDA5NDgsImV4cCI6MjA5MTM3Njk0OH0.Wjs298hA8xAnIQ4IQHEq3eEfaygF-6FY_5I4i2Se3ZY"

DEBOUNCE_SECS = 4    # segundos de espera tras detectar cambio
PERIODIC_SECS = 300  # sincronizacion periodica cada N segundos (0=off)
# ───────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates,return=minimal",
}


def find_excel(folder: Path):
    """Devuelve el .xlsm mas reciente que empiece por EXCEL_PREFIX."""
    matches = sorted(
        folder.glob(f"{EXCEL_PREFIX}*.xlsm"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return matches[0] if matches else None


def _get(pattern, text):
    m = re.search(pattern, str(text), re.IGNORECASE)
    return m.group(1).strip() if m else ""


def parse_excel(path: str) -> list:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Hoja '{SHEET_NAME}' no encontrada. Hojas: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    arts = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        ceco = str(row[0]).strip() if row[0] else ""
        if not ceco or ceco == "None":
            continue
        raw_cod  = str(row[1]) if row[1] else ""
        ref      = str(row[2]).strip() if row[2] else ""
        fab      = str(row[4]).strip() if row[4] else ""
        cod_fab  = str(row[5]).strip() if row[5] else ""
        raw_prov = str(row[6]) if row[6] else ""

        dali = _get(r"DALI[:\s]+0*(\d+)", raw_cod)
        if dali:
            dali = dali.zfill(8)
        sap   = _get(r"SAP[:\s]+(\d+)", raw_cod)
        casad = _get(r"CASA DELFIN[:\s]+([\w\d]+)", raw_prov)
        pilsa = _get(r"PILSA[:\s]+([\w\d]+)", raw_prov)
        uid   = dali if dali else hashlib.md5(ref.encode()).hexdigest()[:12]

        arts.append({
            "id":                    uid,
            "dali":                  dali,
            "sap":                   sap,
            "referencia":            ref,
            "fabricante":            fab,
            "codigo_fabrica":        cod_fab,
            "departamento":          ceco,
            "proveedor_casa_delfin": "SI" if casad else "NO",
            "proveedor_pilsa":       "SI" if pilsa else "NO",
            "codigo_casa_delfin":    casad,
            "codigo_pilsa":          pilsa,
        })
    wb.close()
    return arts


def sync(arts: list) -> bool:
    if not arts:
        log.info("Sin articulos para subir.")
        return False
    errores = 0
    for i in range(0, len(arts), 100):
        bloque = arts[i:i + 100]
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/articulos_dossier",
            json=bloque,
            headers=HEADERS,
            timeout=30,
        )
        if r.ok:
            log.info(f"  OK bloque {i//100 + 1}: {len(bloque)} articulos")
        else:
            log.error(f"  ERROR bloque {i//100 + 1}: {r.status_code} {r.text[:120]}")
            errores += 1
    try:
        requests.post(
            f"{SUPABASE_URL}/rest/v1/dossier_sync_log",
            json={"total": len(arts), "errores": errores,
                  "ts": datetime.now().isoformat()},
            headers={**HEADERS, "Prefer": "return=minimal"},
            timeout=10,
        )
    except Exception:
        pass
    if errores == 0:
        log.info(f"SYNC OK: {len(arts)} articulos subidos")
    else:
        log.warning(f"SYNC con {errores} errores")
    return errores == 0


def sincronizar(path: str):
    log.info(f"Leyendo: {Path(path).name}")
    try:
        arts = parse_excel(path)
        log.info(f"  {len(arts)} articulos parseados")
        sync(arts)
    except PermissionError:
        log.warning("  Excel ocupado, reintentando en 5s...")
        time.sleep(5)
        try:
            arts = parse_excel(path)
            sync(arts)
        except Exception as e:
            log.error(f"  Error: {e}")
    except Exception as e:
        log.error(f"  Error: {e}")


class Handler(FileSystemEventHandler):
    def __init__(self, folder: Path):
        self.folder = folder
        self._last = 0

    def on_modified(self, event):
        if event.is_directory:
            return
        p = Path(event.src_path)
        if p.name.startswith("~$"):
            return
        if not (p.suffix.lower() == ".xlsm" and p.stem.startswith(EXCEL_PREFIX)):
            return
        now = time.time()
        if now - self._last < DEBOUNCE_SECS:
            return
        self._last = now
        log.info(f"CAMBIO detectado: {p.name}")
        time.sleep(DEBOUNCE_SECS)
        ex = find_excel(self.folder) or p
        sincronizar(str(ex))

    on_created = on_modified


def main():
    folder = Path(EXCEL_FOLDER)
    if not folder.exists():
        log.error(f"Carpeta no encontrada:\n  {folder}")
        sys.exit(1)

    excel = find_excel(folder)
    if excel is None:
        log.error(f"No hay .xlsm con prefijo '{EXCEL_PREFIX}' en:\n  {folder}")
        sys.exit(1)

    log.info("=" * 55)
    log.info("  WATCHER DOSSIER PRINCESS")
    log.info("=" * 55)
    log.info(f"  Carpeta : {folder}")
    log.info(f"  Excel   : {excel.name}")
    log.info(f"  Supabase: {SUPABASE_URL}")
    log.info("")

    log.info("Sincronizacion inicial...")
    sincronizar(str(excel))

    handler  = Handler(folder)
    observer = Observer()
    observer.schedule(handler, str(folder), recursive=False)
    observer.start()
    log.info(f"Vigilando cambios en: {folder}")
    log.info("Pulsa Ctrl+C para detener.\n")

    last_periodic = time.time()
    try:
        while True:
            time.sleep(1)
            if PERIODIC_SECS > 0 and (time.time() - last_periodic) >= PERIODIC_SECS:
                log.info("Sincronizacion periodica...")
                ex = find_excel(folder)
                if ex:
                    sincronizar(str(ex))
                last_periodic = time.time()
    except KeyboardInterrupt:
        log.info("Deteniendo...")
        observer.stop()
    observer.join()
    log.info("Watcher detenido.")


if __name__ == "__main__":
    main()
